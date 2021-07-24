import os
import sys
import openpyxl
from openpyxl import Workbook,load_workbook
path='术业专攻\sheet另存为工作簿\拆分前/'
new_path='术业专攻\sheet另存为工作簿\拆分后/'
files=os.listdir(path)
for file in files:
    length=len(file)
    file_path=path+file   
    wb = load_workbook(file_path)
    sheets = wb.sheetnames 
    for s in sheets:
        new_file_path=new_path+file[:(length-5)]+'('+s+')'+'.xlsx'
        wb = load_workbook(file_path)
        for i in sheets:
            if s==i:
                pass
            else:
                wb.remove(wb[i])
        wb.save(new_file_path)
