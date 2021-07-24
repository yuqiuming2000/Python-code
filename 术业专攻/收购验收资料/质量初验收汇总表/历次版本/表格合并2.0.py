import os
import sys
from openpyxl import Workbook,load_workbook
from copy import copy
from openpyxl.descriptors.base import Length
from openpyxl.utils import get_column_letter
path='术业专攻\收购验收资料\质量初验收汇总表\调整前/'
new_path='术业专攻\收购验收资料\质量初验收汇总表\调整后/'
files=os.listdir(path)
new_wb=load_workbook('术业专攻\收购验收资料\质量初验收汇总表\不能删除\质量验收表模板.xlsx')
for file in files:
    file_path=path+file
    new_file_path=new_path+'质量初验收汇总表.xlsx'
    wb=load_workbook(file_path)
    print(file_path)
    ws=wb['附件5-1  质量初验收汇总表']   
    items=['横岗','昌碧','武阳','三江','广福','南新','新联','泾口','塘南','富山','冈上','向塘','塔城','蒋巷','莲塘','幽兰']
    # items_2=['江西南昌横岗国家粮食储备库','江西南昌昌碧米业集团有限公司','南昌县武阳粮食管理所','南昌县三江粮食管理所','南昌县广福粮食管理所','南昌县南新粮食管理所','南昌县新联粮食管理所','南昌县泾口粮食管理所','南昌县塘南粮食管理所','南昌县富山粮食管理所','南昌县冈上粮食管理所','南昌县向塘粮食管理所','南昌县塔城粮食管理所','南昌县蒋巷粮食管理所','南昌县莲塘粮食管理所','南昌县幽兰粮食管理所']
    items_2=['江西南昌横岗国家粮食储备库','江西南昌昌碧米业集团有限公司','南昌县武阳粮食管理所','南昌县三江粮食管理所','南昌县广福粮食管理所','南昌县南新粮食管理所','南昌县新联粮食管理所','南昌县泾口粮食管理所','南昌县塘南粮食管理所','南昌县富山粮食管理所','南昌县冈上粮食管理所','南昌县向塘粮食管理所','南昌县塔城粮食管理所','南昌县蒋巷粮食管理所','南昌县莲塘粮食管理所','南昌县幽兰粮食管理所']
    for item in items:
        if item in file:
            new_ws=new_wb.copy_worksheet(new_wb['模板'])
            new_ws.title=item

            sum=0
            for r in range(6,ws.max_row-1):
 
                for c in range(3,ws.max_column+1):
                    c=get_column_letter(c)
                    m=c+str(r)
                    new_ws['{}'.format(m)].value=ws['{}'.format(m)].value
                for item_2 in items_2:
                    if item in item_2:
                        new_ws['A3'].value='检验单位（盖章）：'+item_2
                    if item_2 in str(new_ws['{}'.format('C'+str(r))].value):
                        s = new_ws['{}'.format('C'+str(r))].value
                        Length=len(item_2)
                        print(s)
                        print(Length)
                        s= s[Length:]
                        new_ws['{}'.format('C'+str(r))].value=s
                    
                    
                if new_ws['{}'.format('D'+str(r))].value!=None:
                    new_ws['{}'.format('A'+str(r))].value='南昌县'
                    new_ws['{}'.format('B'+str(r))].value='南昌直属库'
                    new_ws['{}'.format('W'+str(r))].value='是'


                if new_ws['{}'.format('D'+str(r))].value==None:
                    if r==7:
                        break
                    else:
                        for i in range(6,r):
                            sum=sum+int(new_ws['{}'.format('V'+str(i))].value)
                        Difference=new_ws.max_row-2-r
                        if Difference>1:
                            new_ws['{}'.format('C'+str(r+2))].value='合计'
                            new_ws['{}'.format('V'+str(r+2))].value=sum
                        else:
                            new_ws['{}'.format('C'+str(r))].value='合计'
                            new_ws['{}'.format('V'+str(r))].value=sum
                        break



new_wb.remove(new_wb['模板'])
new_wb.save(new_file_path)  