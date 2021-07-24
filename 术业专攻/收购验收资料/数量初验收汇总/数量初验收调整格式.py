import os
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Alignment,Side,Border, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import re
path='术业专攻\收购验收资料\数量初验收汇总\调整前/'
new_path='术业专攻\收购验收资料\数量初验收汇总\调整后/'
files=os.listdir(path)
for file in files:
    file_path=path+file
    new_file_path=new_path+'修改后'+file
    wb=load_workbook(file_path)
    ws=wb['附件4-1 数量初验收汇总表 ']
    ws.column_dimensions['A'].width=11
    ws.column_dimensions['B'].width=11
    ws.column_dimensions['C'].width=11
    # ws.column_dimensions['D'].width=6.5

    ws.row_dimensions[1].height=18
    ws.row_dimensions[2].height=25
    ws.row_dimensions[3].height=25
    ws.row_dimensions[4].height=27
    ws.row_dimensions[5].height=27
    side=Side('thin')
    for i in range(6,ws.max_row+1):
        ws.row_dimensions[i].height=23
    for i in range(4,ws.max_column+1):
        i=get_column_letter(i)
        ws.column_dimensions[i].width=7.5
    for row in ws.iter_rows():
        for cell in row:
            cell.font=Font(name='宋体',size=10)      
    for row in ws.iter_rows(min_row=4,max_row=ws.max_row-2):
        for cell in row:
            cell.font=Font(name='宋体',size=10)
            cell.fill=PatternFill('solid',fgColor='FFFFFFFF')
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            cell.border=Border(top=side,bottom=side,right=side,left=side)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
    ws.merge_cells('I4:M4')
    ws.merge_cells('N4:N5')
    ws.merge_cells('O4:O5')
    ws.merge_cells('P4:P5')
    ws.merge_cells('A3:F3')

 
    ws['A1'].font=Font(name='宋体',size=10,bold=True)
    for cell in ws[2]:
        cell.alignment=Alignment(horizontal='center',vertical='center')
        cell.font=Font(name='宋体',size=16,bold=True)
    col_list=['E','F','G','H','P']
    for col in col_list:
        for cell in ws[col]:
            cell.number_format = '0_);[Red](0)'

    for cell in ws[ws.max_row]:
        cell.alignment=Alignment(horizontal='left',vertical='center')
        cell.font=Font(name='宋体',size=10)

    #以下为设置打印
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = '1:5' 
    ws.page_margins = PageMargins(top=0.4, bottom=0.4,left=0.2,right=0.2)#英寸需要换算成毫米
    wb.save(new_file_path)
