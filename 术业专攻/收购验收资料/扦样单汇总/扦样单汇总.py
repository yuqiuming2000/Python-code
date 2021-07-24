import os
import sys
from openpyxl import Workbook,load_workbook
from copy import copy
from openpyxl.descriptors.base import Length
from openpyxl.utils import get_column_letter
path='术业专攻\收购验收资料\扦样单汇总\调整前/'
new_path='术业专攻\收购验收资料\扦样单汇总\调整后/'
files=os.listdir(path)
new_wb=load_workbook('术业专攻\收购验收资料\扦样单汇总\扦样单模板.xlsx')
f_ws=new_wb.copy_worksheet(new_wb['模板'])
f_ws.title='汇总'
number=1000
for file in files:
    file_path=path+file
    new_file_path=new_path+'扦样单汇总表.xlsx'
    wb=load_workbook(file_path)
    print(file_path)
    ws=wb['附件3  扦样单']   
    items=['横岗','昌碧','武阳','三江','广福','南新','新联','泾口','塘南','富山','冈上','向塘','塔城','蒋巷','莲塘','幽兰']
    # items_2=['江西南昌横岗国家粮食储备库','江西南昌昌碧米业集团有限公司','南昌县武阳粮食管理所','南昌县三江粮食管理所','南昌县广福粮食管理所','南昌县南新粮食管理所','南昌县新联粮食管理所','南昌县泾口粮食管理所','南昌县塘南粮食管理所','南昌县富山粮食管理所','南昌县冈上粮食管理所','南昌县向塘粮食管理所','南昌县塔城粮食管理所','南昌县蒋巷粮食管理所','南昌县莲塘粮食管理所','南昌县幽兰粮食管理所']
    items_2=['江西南昌横岗国家粮食储备库','江西南昌昌碧米业集团有限公司','南昌县武阳粮食管理所','南昌县三江粮食管理所','南昌县广福粮食管理所','南昌县南新粮食管理所','南昌县新联粮食管理所','南昌县泾口粮食管理所','南昌县塘南粮食管理所','南昌县富山粮食管理所','南昌县冈上粮食管理所','南昌县向塘粮食管理所','南昌县塔城粮食管理所','南昌县蒋巷粮食管理所','南昌县莲塘粮食管理所','南昌县幽兰粮食管理所']
    
    for item in items:
        if item in file:
            new_ws=new_wb.copy_worksheet(new_wb['模板'])
            new_ws.title=item
            sum=0           
            g_list=['E5','B7','B8']
            for g in g_list:
                new_ws[g].value=ws[g].value
         
            for r in range(10,ws.max_row-7):
                if ws['{}'.format('D'+str(r))].value==None:
                    print(r)
                    for i in range(10,r):
                        for c in range(2,ws.max_column+1):
                            c=get_column_letter(c)
                            new_ws['{}'.format(c+str(i))].value=ws['{}'.format(c+str(i))].value
                        sum=sum+int(new_ws['{}'.format('F'+str(i))].value)
                        number+=1
                        new_ws['{}'.format('D'+str(i))].value='散装'
                        new_ws['{}'.format('A'+str(i))].value='NCXR'+str(number)
                                  
#以上部分复制正文内容
                        for item_2 in items_2:
                            if item in item_2:
                                new_ws['B6'].value=item_2
                            if item_2 in str(new_ws['{}'.format('C'+str(i))].value):
                                s = new_ws['{}'.format('B'+str(i))].value
                                Length=len(item_2)
                                print(s)
                                print(Length)
                                s= s[Length:]
                                new_ws['{}'.format('C'+str(i))].value=s
                                #以上内容检查A3检验单位和C4切片切掉粮管所名
                    for row in new_ws.iter_rows(min_row=10,max_row=r-1,values_only=True):
                        f_ws.append(row)
                        #以上内容复制单表内容到汇总表
                    print(new_ws.max_row)
                    if r>11:
                        Difference=new_ws.max_row-7-r
                        if Difference>1:
                            new_ws['{}'.format('C'+str(r+2))].value='合计'
                            new_ws['{}'.format('F'+str(r+2))].value=sum

                        else:
                            new_ws['{}'.format('C'+str(r))].value='合计'
                            new_ws['{}'.format('F'+str(r))].value=sum
                    #以上内容判断空行的位置，填写合计和求和    
                    break
new_wb.remove(new_wb['模板'])
new_wb.save(new_file_path)  