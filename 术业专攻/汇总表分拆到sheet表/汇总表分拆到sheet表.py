from openpyxl import Workbook,load_workbook
wb=load_workbook('术业专攻\拆分表格\春季普查质量自查汇总表.xlsx')
ws=wb.active
# items=['江西南昌横岗国家粮食储备库','江西南昌昌碧米业集团有限公司','南昌县武阳粮食管理所','南昌县三江粮食管理所','南昌县广福粮食管理所','南昌县南新粮食管理所','南昌县新联粮食管理所','南昌县泾口粮食管理所','南昌县塘南粮食管理所','南昌县富山粮食管理所','南昌县冈上粮食管理所','南昌县向塘粮食管理所','南昌县塔城粮食管理所','南昌县蒋巷粮食管理所','南昌县莲塘粮食管理所','南昌县幽兰粮食管理所']
items=['江西南昌横岗国家粮食储备库','江西南昌昌碧米业集团有限公司','南昌县武阳粮食管理所','南昌县三江粮食管理所','南昌县广福粮食管理所','南昌县南新粮食管理所','南昌县新联粮食管理所','南昌县泾口粮食管理所','南昌县塘南粮食管理所','南昌县富山粮食管理所','南昌县冈上粮食管理所','南昌县向塘粮食管理所','南昌县塔城粮食管理所','南昌县蒋巷粮食管理所','南昌县莲塘粮食管理所','南昌县幽兰粮食管理所']
for item in items:
    path='术业专攻\拆分表格\拆分后/'
    file_path=path+item+'.xlsx'
    print(file_path)
    new_wb=Workbook()
    new_ws=new_wb.active

    for row in ws.iter_rows(max_row=4,values_only=True):
        new_ws.append(row)
    for row in ws.iter_rows(min_row=5,values_only=True):
        if item in row[2]:
            new_ws.append(row)
    new_ws['A3'].value='自查单位：（盖章）'+item
    new_wb.save(file_path)  
   

