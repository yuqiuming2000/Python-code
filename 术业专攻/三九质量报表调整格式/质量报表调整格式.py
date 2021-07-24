import os
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Alignment,Side,Border, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
path='术业专攻\三九质量报表调整格式\调整前/'
new_path='术业专攻\三九质量报表调整格式\调整后/'
files=os.listdir(path)
for file in files:
    file_path=path+file
    new_file_path=new_path+file
    wb=load_workbook(file_path)
    ws=wb.active
    ws.column_dimensions['A'].width=6.5
    ws.column_dimensions['B'].width=6.5
    ws.column_dimensions['C'].width=35
    ws.column_dimensions['D'].width=8.5
    ws.column_dimensions['E'].width=9
    ws.column_dimensions['F'].width=9
    ws.column_dimensions['G'].width=8.5
    ws.column_dimensions['H'].width=9.5
    ws.row_dimensions[1].height=18
    ws.row_dimensions[2].height=25
    ws.row_dimensions[3].height=18
    ws.row_dimensions[4].height=70
   
    for i in range(5,ws.max_row+1):
        ws.row_dimensions[i].height=24
    for i in range(9,ws.max_column+1):
        i=get_column_letter(i)
        ws.column_dimensions[i].width=6.0
    ws.column_dimensions['T'].width=7
    ws.column_dimensions['O'].width=5
    side=Side('thin')
    for cell in ws[4]:
        cell.border=Border(top=side,bottom=side,right=side,left=side)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.font=Font(name='宋体',size=12,bold=True)
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.font=Font(name='宋体',size=11)
            cell.fill=PatternFill('solid',fgColor='FFFFFFFF')
            cell.alignment=Alignment(horizontal='center',vertical='center')
            cell.border=Border(top=side,bottom=side,right=side,left=side)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
    for cell in ws[2]:
        # cell.border=Border(top=side,bottom=side,right=side,left=side)
        cell.alignment=Alignment(horizontal='center',vertical='center')
        cell.font=Font(name='宋体',size=16,bold=True)

    #以下为设置文本小数点位数    
    for cell in ws['S']:
        cell.number_format = '0_);[Red](0)'
    col_list=['I','J','K','L','M','N','R']
    for col in col_list:
        for cell in ws[col]:
            if cell.value==0:
                cell.number_format = '0_);[Red](0)'
            else:
                cell.number_format = '0.0'
    for r in range (5,ws.max_row+1):
        cell_T='T'+str(r)
      
        if ws[cell_T].value>=0.1:
            ws[cell_T].number_format = '0.00'
        else:
            ws[cell_T].number_format = '0.000'

   #以下为设置打印
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = '1:4' 
    ws.page_margins = PageMargins(top=0.4, bottom=0.4,left=0.0038,right=0.0038)#英寸需要换算成毫米
    ws.orientation='landscape'

    wb.save(new_file_path)

