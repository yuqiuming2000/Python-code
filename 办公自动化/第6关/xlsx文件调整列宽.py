from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Alignment,Side,Border, alignment
content_fill=PatternFill('solid',fgColor='FF7F24')
align=Alignment(horizontal='left',vertical='center')
header_align=Alignment(horizontal='center',vertical='center')
side=Side('thin')
header_border=Border(bottom=side,right=side)
content_border=Border(top=side,bottom=side,right=side,left=side)
wb=load_workbook('当当网2021年6月份销售排行榜.xlsx')
ws=wb.active
ws.column_dimensions['A'].width=5
ws.column_dimensions['B'].width=62
ws.column_dimensions['C'].width=40
ws.column_dimensions['D'].width=45
ws.column_dimensions['E'].width=15
ws.column_dimensions['F'].width=10
ws.column_dimensions['G'].width=10
for cell in ws[1]:
    cell.border=header_border
    cell.alignment=header_align
for row in ws.iter_rows(min_row=2):
    for cell in row:
        # cell.fill=content_fill
        cell.alignment=align
        cell.border=content_border
wb.save('当当网2021年6月份销售排行榜修改后.xlsx')
